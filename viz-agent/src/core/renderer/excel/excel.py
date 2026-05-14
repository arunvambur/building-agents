import os
import re
import tempfile
import uuid
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.dsl.schema import Chart, VisualizationSpec
from core.renderer.base import Renderer

_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")

# Chart types that map to openpyxl BarChart with type="bar" (horizontal)
_HORIZONTAL_TYPES = {"horizontal_bar"}
# Chart types that map to openpyxl BarChart with grouping="stacked"
_STACKED_TYPES = {"stacked_bar"}
# Chart types that map to openpyxl BarChart with grouping="clustered" (side-by-side)
_GROUPED_TYPES = {"grouped_bar"}
# Chart types that map to openpyxl AreaChart
_AREA_TYPES = {"area"}
# Donut maps to PieChart with hole
_DONUT_TYPES = {"donut"}
# Histogram maps to BarChart (binned data)
_HISTOGRAM_TYPES = {"histogram"}


class ExcelRenderer(Renderer):

    name = "excel"

    def supports(self, format: str) -> bool:
        return format.lower() == "excel"

    def render(self, spec: VisualizationSpec, data: Any) -> str:
        """
        Renders a VisualizationSpec to an Excel workbook (.xlsx).
        Returns a 'file://<path>' string for API-layer detection.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        rows: list[dict] = data if isinstance(data, list) else []

        for i, chart_spec in enumerate(spec.charts):
            sheet_title = self._sheet_title(chart_spec.title, i)
            ws = wb.create_sheet(title=sheet_title)
            self._write_data_sheet(ws, rows, chart_spec)
            self._add_chart(ws, rows, chart_spec)

        if spec.filters:
            self._write_filters_sheet(wb, spec)

        output_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}.xlsx")
        wb.save(output_path)
        return f"file://{output_path}"

    # ---- internals ----

    def _sheet_title(self, title: str | None, index: int) -> str:
        title = title or f"Chart {index + 1}"
        title = _INVALID_SHEET_TITLE_CHARS.sub("-", title).strip()
        return (title or f"Chart {index + 1}")[:31]

    def _write_data_sheet(self, ws, rows: list[dict], chart_spec: Chart) -> None:
        if not rows:
            ws.append(["No data available"])
            return

        headers = list(rows[0].keys())
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([row.get(h) for h in headers])

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(header)), *(len(str(r.get(header, ""))) for r in rows))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def _add_chart(self, ws, rows: list[dict], chart_spec: Chart) -> None:
        if not rows:
            return

        headers = list(rows[0].keys())
        x_field = chart_spec.x.field
        y_field = chart_spec.y.field

        if x_field not in headers or y_field not in headers:
            return

        x_col = headers.index(x_field) + 1
        y_col = headers.index(y_field) + 1
        y2_col = None
        if chart_spec.y2 and chart_spec.y2.field in headers:
            y2_col = headers.index(chart_spec.y2.field) + 1

        data_rows = len(rows)
        chart = self._build_chart(chart_spec, ws, x_col, y_col, y2_col, data_rows)
        if chart is None:
            return

        chart.title = chart_spec.title or f"{y_field} by {x_field}"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        anchor_row = data_rows + 4
        ws.add_chart(chart, f"A{anchor_row}")

    def _build_chart(
        self, chart_spec: Chart, ws, x_col: int, y_col: int,
        y2_col: int | None, data_rows: int
    ):
        data_ref = Reference(ws, min_col=y_col, min_row=1, max_row=data_rows + 1)
        cats_ref = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
        chart_type = chart_spec.type

        if chart_type == "bar":
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "horizontal_bar":
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "clustered"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "stacked_bar":
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.add_data(data_ref, titles_from_data=True)
            if y2_col:
                chart.add_data(Reference(ws, min_col=y2_col, min_row=1, max_row=data_rows + 1),
                               titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "grouped_bar":
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.add_data(data_ref, titles_from_data=True)
            if y2_col:
                chart.add_data(Reference(ws, min_col=y2_col, min_row=1, max_row=data_rows + 1),
                               titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "line":
            chart = LineChart()
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "area":
            from openpyxl.chart import AreaChart
            chart = AreaChart()
            chart.grouping = "standard"
            chart.add_data(data_ref, titles_from_data=True)
            if y2_col:
                chart.add_data(Reference(ws, min_col=y2_col, min_row=1, max_row=data_rows + 1),
                               titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "pie":
            chart = PieChart()
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "donut":
            from openpyxl.chart import DoughnutChart
            chart = DoughnutChart()
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "histogram":
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        if chart_type == "scatter":
            from openpyxl.chart import ScatterChart, Series
            chart = ScatterChart()
            x_ref = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
            y_ref = Reference(ws, min_col=y_col, min_row=1, max_row=data_rows + 1)
            series = Series(y_ref, x_ref, title_from_data=True)
            chart.series.append(series)
            return chart

        if chart_type == "bubble":
            from openpyxl.chart import BubbleChart, Series
            chart = BubbleChart()
            x_ref = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
            y_ref = Reference(ws, min_col=y_col, min_row=2, max_row=data_rows + 1)
            size_ref = Reference(ws, min_col=y2_col if y2_col else y_col,
                                 min_row=2, max_row=data_rows + 1)
            series = Series(values=y_ref, xvalues=x_ref, zvalues=size_ref)
            chart.series.append(series)
            return chart

        if chart_type == "waterfall":
            # Approximated as stacked bar in Excel (no native waterfall in openpyxl)
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            return chart

        # heatmap and gauge have no native Excel chart equivalent —
        # the data table written by _write_data_sheet serves as the output.
        if chart_type in ("heatmap", "gauge"):
            return None

        return None


    def _write_filters_sheet(self, wb: openpyxl.Workbook, spec: VisualizationSpec) -> None:
        ws = wb.create_sheet(title="Filters Applied")
        ws.append(["Field", "Operator", "Value"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for f in spec.filters:
            ws.append([f.field, f.op, f.value])
