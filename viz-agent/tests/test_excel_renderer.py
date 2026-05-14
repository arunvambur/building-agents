import os

import pytest

from core.dsl.schema import Aggregation, Axis, Chart, Filter, VisualizationSpec
from core.renderer.excel.excel import ExcelRenderer

FILE_PREFIX = "file://"


def strip_prefix(result: str) -> str:
    assert result.startswith(FILE_PREFIX), f"Expected file:// prefix, got: {result[:30]}"
    return result[len(FILE_PREFIX):]


@pytest.fixture
def renderer():
    return ExcelRenderer()


@pytest.fixture
def sample_data():
    return [
        {"town": "St Ives",  "rating": 4.8, "price_single": 140.0, "price_double": 210.0, "available_rooms": 6},
        {"town": "Newquay",  "rating": 4.5, "price_single": 120.0, "price_double": 180.0, "available_rooms": 5},
        {"town": "Falmouth", "rating": 4.2, "price_single": 95.0,  "price_double": 150.0, "available_rooms": 2},
    ]


def _make_spec(chart_type, output="excel", aggregation=None, y2_field=None, title=None):
    return VisualizationSpec(
        charts=[Chart(
            type=chart_type,
            x=Axis(field="town", type="dimension"),
            y=Axis(field="rating", type="measure"),
            y2=Axis(field=y2_field, type="measure") if y2_field else None,
            aggregation=aggregation,
            title=title or f"{chart_type.replace('_', ' ').title()} Chart",
        )],
        output=output,
    )


# ---- supports ----

def test_supports_excel(renderer):
    assert renderer.supports("excel") is True

def test_does_not_support_tableau(renderer):
    assert renderer.supports("tableau") is False


# ---- original chart types ----

def test_render_returns_file_prefix(renderer, sample_data):
    result = renderer.render(_make_spec("bar"), sample_data)
    assert result.startswith(FILE_PREFIX)
    path = strip_prefix(result)
    assert path.endswith(".xlsx")
    assert os.path.exists(path)
    os.remove(path)

def test_render_creates_correct_sheet_name(renderer, sample_data):
    import openpyxl
    path = strip_prefix(renderer.render(_make_spec("bar", title="Avg Rating by Town"), sample_data))
    wb = openpyxl.load_workbook(path)
    assert "Avg Rating by Town" in wb.sheetnames
    os.remove(path)

def test_render_sanitizes_invalid_sheet_name_characters(renderer, sample_data):
    import openpyxl
    path = strip_prefix(renderer.render(_make_spec("bar", title="Rating: [Town]"), sample_data))
    wb = openpyxl.load_workbook(path)
    assert any("Rating" in s for s in wb.sheetnames)
    os.remove(path)

def test_render_with_filters_creates_filter_sheet(renderer, sample_data):
    import openpyxl
    spec = VisualizationSpec(
        charts=[Chart(type="bar", x=Axis(field="town", type="dimension"),
                      y=Axis(field="rating", type="measure"), title="Filtered")],
        filters=[Filter(field="rating", op=">", value="4.0")],
        output="excel",
    )
    path = strip_prefix(renderer.render(spec, sample_data))
    wb = openpyxl.load_workbook(path)
    assert "Filters Applied" in wb.sheetnames
    os.remove(path)

def test_render_with_empty_data(renderer):
    path = strip_prefix(renderer.render(_make_spec("bar"), []))
    assert os.path.exists(path)
    os.remove(path)

def test_render_line_chart(renderer, sample_data):
    path = strip_prefix(renderer.render(_make_spec("line"), sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_pie_chart(renderer, sample_data):
    spec = _make_spec("pie", aggregation=Aggregation(field="available_rooms", op="sum"))
    path = strip_prefix(renderer.render(spec, sample_data))
    assert os.path.exists(path)
    os.remove(path)


# ---- new chart types ----

def test_render_horizontal_bar(renderer, sample_data):
    path = strip_prefix(renderer.render(_make_spec("horizontal_bar"), sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_stacked_bar(renderer, sample_data):
    spec = _make_spec("stacked_bar", y2_field="price_single")
    path = strip_prefix(renderer.render(spec, sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_grouped_bar(renderer, sample_data):
    spec = _make_spec("grouped_bar", y2_field="price_single")
    path = strip_prefix(renderer.render(spec, sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_area_chart(renderer, sample_data):
    path = strip_prefix(renderer.render(_make_spec("area"), sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_area_chart_with_y2(renderer, sample_data):
    spec = _make_spec("area", y2_field="price_single")
    path = strip_prefix(renderer.render(spec, sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_donut_chart(renderer, sample_data):
    spec = _make_spec("donut", aggregation=Aggregation(field="available_rooms", op="sum"))
    path = strip_prefix(renderer.render(spec, sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_histogram(renderer, sample_data):
    path = strip_prefix(renderer.render(_make_spec("histogram"), sample_data))
    assert os.path.exists(path)
    os.remove(path)

def test_render_scatter_chart(renderer, sample_data):
    path = strip_prefix(renderer.render(_make_spec("scatter"), sample_data))
    assert os.path.exists(path)
    os.remove(path)
